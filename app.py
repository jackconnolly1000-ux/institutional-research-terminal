import os
import json
import requests
import re
import base64
import time
import concurrent.futures
import openpyxl
import yfinance as yf
import pandas as pd
import plotly.graph_objects as go
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
from datetime import datetime
from bs4 import BeautifulSoup
from anthropic import Anthropic
import streamlit as st
from st_keyup import st_keyup

# --- STREAMLIT PAGE CONFIG & PROFESSIONAL BLOOMBERG/FACTSET CSS ---
st.set_page_config(
    page_title="Institutional Research Terminal v9.6",
    layout="wide",
    initial_sidebar_state="collapsed"
)

st.markdown("""
    <style>
    /* Global High-Contrast Professional Theme */
    .stApp, .stMarkdown, p, span, label, div {
        color: #d1d4dc !important;
        font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, Helvetica, Arial, sans-serif;
    }
    .stApp {
        background-color: #131722 !important;
    }
    
    /* Typography */
    h1 { color: #ffffff !important; font-weight: 700; font-size: 1.8rem !important; letter-spacing: -0.5px; }
    h2, h3, h4 { color: #f8fafc !important; font-weight: 600; }
    .stCaption, small { color: #787b86 !important; }
    
    /* Completely Neutralize Container Block Backgrounds & Borders */
    div[data-testid="stHorizontalBlock"],
    div[data-testid="stVerticalBlock"],
    div[data-testid="column"],
    div[data-testid="element-container"],
    div[data-testid="stToolbar"] {
        background-color: transparent !important;
        border: none !important;
        box-shadow: none !important;
    }
    
    /* Target st_keyup Component Iframe to Match Dark Theme */
    iframe[title="st_keyup.st_keyup"] {
        background-color: #1e222d !important;
        border-radius: 4px !important;
        border: 1px solid #363c4e !important;
    }
    
    /* Flush Command Toolbar Container */
    .terminal-toolbar {
        background-color: #1e222d;
        border: 1px solid #2a2e39;
        padding: 16px;
        border-radius: 6px;
        margin-bottom: 10px;
    }
    
    /* Autocomplete Dropdown Container */
    .autocomplete-dropdown {
        background-color: #181c25;
        border: 1px solid #2a2e39;
        border-top: none;
        border-radius: 0 0 6px 6px;
        padding: 12px 16px;
        margin-bottom: 20px;
        box-shadow: 0 8px 20px rgba(0,0,0,0.5);
    }
    
    /* Professional Action Button */
    .stButton button {
        background: #2962ff !important;
        color: white !important;
        font-weight: 600;
        border: none;
        border-radius: 4px;
        padding: 0.45rem 0.9rem;
        font-size: 0.9rem;
        transition: background 0.2s ease;
        margin-top: 28px;
    }
    .stButton button:hover {
        background: #1e53e5 !important;
    }
    
    /* Metric Cards */
    div[data-testid="stMetric"] {
        background-color: #1e222d;
        border: 1px solid #2a2e39;
        padding: 10px 14px;
        border-radius: 4px;
    }
    div[data-testid="stMetric"] label {
        color: #787b86 !important;
        font-size: 0.75rem !important;
        text-transform: uppercase;
        letter-spacing: 0.5px;
    }
    div[data-testid="stMetric"] div[data-testid="stMetricValue"] {
        color: #ffffff !important;
        font-size: 1.25rem !important;
        font-weight: 600;
        font-family: monospace;
    }
    
    /* Info/Warning Callouts */
    .stAlert {
        background-color: #1e222d !important;
        border: 1px solid #2a2e39 !important;
        color: #d1d4dc !important;
        border-radius: 4px;
    }
    </style>
""", unsafe_allow_html=True)

# --- SECURE KEY RETRIEVAL ---
try:
    if "ANTHROPIC_API_KEY" not in st.secrets:
        st.error("ANTHROPIC_API_KEY was not found in Streamlit Secrets.")
        st.stop()
    raw_key = st.secrets["ANTHROPIC_API_KEY"]
    clean_key = str(raw_key).strip().replace("\n", "").replace(" ", "").replace("\r", "")
    client = Anthropic(api_key=clean_key, timeout=60.0)
except Exception as e:
    st.error(f"Initialization Error: {e}")
    st.stop()

# --- BACKEND FUNCTIONS ---
@st.cache_data(ttl=86400)
def load_sec_directory() -> pd.DataFrame:
    headers = {'User-Agent': 'InstitutionalResearchDirectory admin@research-terminal.com'}
    try:
        resp = requests.get("https://www.sec.gov/files/company_tickers.json", headers=headers, timeout=5)
        if resp.status_code == 200:
            data = resp.json()
            rows = [{"ticker": v['ticker'].upper(), "name": v['title'].title()} for v in data.values()]
            return pd.DataFrame(rows)
    except Exception:
        pass
    return pd.DataFrame([
        {"ticker": "AAPL", "name": "Apple Inc."},
        {"ticker": "MSFT", "name": "Microsoft Corp."},
        {"ticker": "NVDA", "name": "Nvidia Corp."},
        {"ticker": "FAST", "name": "Fastenal Co."},
        {"ticker": "STRL", "name": "Sterling Infrastructure Inc."},
        {"ticker": "AAP", "name": "Advance Auto Parts Inc."}
    ], columns=["ticker", "name"])

@st.cache_data(ttl=3600)
def get_comprehensive_financials(ticker: str) -> dict:
    try:
        tkr = yf.Ticker(ticker)
        info = tkr.info
        mcap = info.get('marketCap', 0)
        mcap_str = f"${mcap/1e12:.2f}T" if mcap >= 1e12 else (f"${mcap/1e9:.2f}B" if mcap >= 1e9 else (f"${mcap/1e6:.2f}M" if mcap > 0 else "N/A"))
        price = info.get('currentPrice') or info.get('regularMarketPrice')
        price_str = f"${price:.2f}" if price else "N/A"
        fpe = info.get('forwardPE')
        fpe_str = f"{fpe:.1f}x" if fpe else "N/A"
        rev_growth = info.get('revenueGrowth')
        rev_str = f"{rev_growth * 100:.1f}%" if rev_growth else "N/A"
        gross_margin = f"{info.get('grossMargins', 0) * 100:.2f}%" if info.get('grossMargins') else "N/A"
        op_margin = f"{info.get('operatingMargins', 0) * 100:.2f}%" if info.get('operatingMargins') else "N/A"
        profit_margin = f"{info.get('profitMargins', 0) * 100:.2f}%" if info.get('profitMargins') else "N/A"
        roe = f"{info.get('returnOnEquity', 0) * 100:.2f}%" if info.get('returnOnEquity') else "N/A"
        debt_to_equity = f"{info.get('debtToEquity', 0):.2f}" if info.get('debtToEquity') else "N/A"
        fcf = info.get('freeCashflow', 0)
        fcf_str = f"${fcf/1e9:.2f}B" if fcf and abs(fcf) >= 1e9 else (f"${fcf/1e6:.2f}M" if fcf else "N/A")
        ebitda = info.get('ebitda', 0)
        ebitda_str = f"${ebitda/1e9:.2f}B" if ebitda and abs(ebitda) >= 1e9 else (f"${ebitda/1e6:.2f}M" if ebitda else "N/A")
        return {"mcap": mcap_str, "price": price_str, "fpe": fpe_str, "rev_growth": rev_str, "gross_margin": gross_margin, "op_margin": op_margin, "profit_margin": profit_margin, "roe": roe, "debt_to_equity": debt_to_equity, "fcf": fcf_str, "ebitda": ebitda_str}
    except Exception as e:
        return {k: "N/A" for k in ["mcap", "price", "fpe", "rev_growth", "gross_margin", "op_margin", "profit_margin", "roe", "debt_to_equity", "fcf", "ebitda"]}

def sec_get_request(url: str, headers: dict, retries: int = 3):
    for attempt in range(retries):
        try:
            resp = requests.get(url, headers=headers, timeout=12)
            if resp.status_code == 200: return resp
            elif resp.status_code == 429: time.sleep(1.5 * (attempt + 1))
        except Exception: time.sleep(1.0)
    return None

@st.cache_data(ttl=3600)
def fetch_firm_data(ticker: str) -> dict:
    headers = {'User-Agent': f'InstitutionalResearch analyst_{ticker.lower()}@research-terminal.com'}
    tickers_url = "https://www.sec.gov/files/company_tickers.json"
    
    resp = sec_get_request(tickers_url, headers)
    if not resp: return None
    try:
        data = resp.json()
    except json.JSONDecodeError:
        return None
        
    cik_padded = next((str(v['cik_str']).zfill(10) for v in data.values() if v['ticker'].upper() == ticker.upper()), None)
    if not cik_padded: return None
        
    subs_url = f"https://data.sec.gov/submissions/CIK{cik_padded}.json"
    time.sleep(0.15)
    subs_resp = sec_get_request(subs_url, headers)
    if not subs_resp: return None
    subs_data = subs_resp.json()
    filings = subs_data['filings']['recent']
    
    company_name = subs_data.get('name', ticker.upper()).title()
    industry_desc = subs_data.get('sicDescription', 'Public Issuer').title()
    form_list = filings['form']
    
    form_4_dates = [filings['filingDate'][i] for i, f in enumerate(form_list) if f in ['4', '4/A']]
    recent_4s = [d for d in form_4_dates if (datetime.now() - datetime.strptime(d, '%Y-%m-%d')).days <= 90]
    insider_velocity = len(recent_4s)
    
    idx_curr = next((i for i, f in enumerate(form_list) if f in ['10-K', '10-Q']), -1)
    if idx_curr == -1: return None
        
    form_type = form_list[idx_curr]
    acc_curr = filings['accessionNumber'][idx_curr].replace('-', '')
    doc_curr = filings['primaryDocument'][idx_curr]
    date_curr = filings['filingDate'][idx_curr]
    cik_no_zeros = str(int(cik_padded))
    
    idx_prior = next((i for i in range(idx_curr + 1, len(form_list)) if form_list[i] in ['10-K', '10-Q']), -1)
    
    risk_regex = r'(?:item\s+1a\b[\.\:\s\-\—\xa0]*(?:risk\s+factors)?)(.*?)(?:item\s+(?:1b|2)\b[\.\:\s\-\—\xa0]|$)'
    mda_regex = r'(?:item\s+7\b[\.\:\s\-\—\xa0]*(?:management[\'\’]?s\s+discussion)?)(.*?)(?:item\s+(?:7a|8)\b[\.\:\s\-\—\xa0]|$)' if form_type == '10-K' else r'(?:item\s+2\b[\.\:\s\-\—\xa0]*(?:management[\'\’]?s\s+discussion)?)(.*?)(?:item\s+(?:3|4)\b[\.\:\s\-\—\xa0]|$)'
    
    url_curr = f"https://www.sec.gov/Archives/edgar/data/{cik_no_zeros}/{acc_curr}/{doc_curr}"
    time.sleep(0.15)
    resp_curr = sec_get_request(url_curr, headers)
    html_curr = resp_curr.content if resp_curr else b""
    text_clean_curr = re.sub(r'\s+', ' ', BeautifulSoup(html_curr, 'html.parser').get_text(separator=' ', strip=True))
    
    risk_raw_curr = max([m.group(1) for m in re.compile(risk_regex, re.I | re.S).finditer(text_clean_curr) if len(m.group(1).strip()) > 4000], key=len, default=text_clean_curr[5000:35000])
    mda_raw_curr = max([m.group(1) for m in re.compile(mda_regex, re.I | re.S).finditer(text_clean_curr) if len(m.group(1).strip()) > 4000], key=len, default=text_clean_curr[40000:80000])
    
    prior_risk_raw = "PRIOR YEAR FILING NOT AVAILABLE"
    if idx_prior != -1:
        try:
            acc_prior = filings['accessionNumber'][idx_prior].replace('-', '')
            doc_prior = filings['primaryDocument'][idx_prior]
            time.sleep(0.15)
            resp_prior = sec_get_request(f"https://www.sec.gov/Archives/edgar/data/{cik_no_zeros}/{acc_prior}/{doc_prior}", headers)
            if resp_prior:
                text_clean_prior = re.sub(r'\s+', ' ', BeautifulSoup(resp_prior.content, 'html.parser').get_text(separator=' ', strip=True))
                prior_risk_raw = max([m.group(1) for m in re.compile(risk_regex, re.I | re.S).finditer(text_clean_prior) if len(m.group(1).strip()) > 4000], key=len, default=text_clean_prior[5000:35000])[:18000]
        except Exception: pass

    eight_k_texts, eight_k_dates, k8_count = "", [], 0
    for idx_form, f_type in enumerate(form_list):
        if f_type in ['8-K', '8-K/A'] and k8_count < 4:
            acc_8k = filings['accessionNumber'][idx_form].replace('-', '')
            doc_8k = filings['primaryDocument'][idx_form]
            date_8k = filings['filingDate'][idx_form]
            time.sleep(0.2)
            resp_8k = sec_get_request(f"https://www.sec.gov/Archives/edgar/data/{cik_no_zeros}/{acc_8k}/{doc_8k}", headers)
            if resp_8k and resp_8k.status_code == 200:
                text_clean_8k = re.sub(r'\s+', ' ', BeautifulSoup(resp_8k.content, 'html.parser').get_text(separator=' ', strip=True))
                if len(text_clean_8k) > 100:
                    eight_k_texts += f"--- 8-K FILED ON {date_8k} ---\n{text_clean_8k[:4500]}\n\n"
                    eight_k_dates.append((date_8k, '8-K'))
                    k8_count += 1
            
    if not eight_k_texts.strip(): eight_k_texts = "NO RECENT 8-K FILINGS IDENTIFIED."
    financial_stats = get_comprehensive_financials(ticker.upper())

    return {
        "ticker": ticker.upper(), "company_name": company_name, "industry": industry_desc, "form_type": form_type,
        "risk": risk_raw_curr.strip()[:20000], "mda": mda_raw_curr.strip()[:20000], "prior_risk": prior_risk_raw.strip(),
        "eight_k": eight_k_texts, "cik": cik_no_zeros, "financials": financial_stats,
        "insider_velocity": insider_velocity, "event_markers": [(date_curr, form_type)] + eight_k_dates
    }

UNIFIED_EXTRACTION_TOOL = {
    "name": "record_comprehensive_firm_analysis",
    "description": "Record catalysts, vulnerabilities, recent 8-K material events, and executive Q&A.",
    "input_schema": {
        "type": "object",
        "properties": {
            "bull_items": {
                "type": "array",
                "items": {"type": "object", "properties": {"target": {"type": "string"}, "category": {"type": "string"}, "description": {"type": "string"}, "impact": {"type": "string", "enum": ["HIGH", "MEDIUM"]}, "quote": {"type": "string"}}, "required": ["target", "category", "description", "impact", "quote"]}
            },
            "bear_items": {
                "type": "array",
                "items": {"type": "object", "properties": {"target": {"type": "string"}, "category": {"type": "string"}, "description": {"type": "string"}, "impact": {"type": "string", "enum": ["HIGH", "MEDIUM"]}, "trend": {"type": "string", "enum": ["NEW", "ESCALATED", "STABLE"]}, "quote": {"type": "string"}}, "required": ["target", "category", "description", "impact", "trend", "quote"]}
            },
            "eight_k_insights": {
                "type": "array",
                "items": {"type": "object", "properties": {"event_date": {"type": "string"}, "category": {"type": "string", "enum": ["LEADERSHIP", "M&A / CONTRACT", "FINANCIAL", "COMPLIANCE", "OPERATIONAL"]}, "takeaway": {"type": "string"}, "quote": {"type": "string"}}, "required": ["event_date", "category", "takeaway", "quote"]}
            },
            "strategic_questions": {"type": "array", "items": {"type": "string"}}
        },
        "required": ["bull_items", "bear_items", "eight_k_insights", "strategic_questions"]
    }
}

@st.cache_data(ttl=3600)
def extract_unified_profile(sec_payload: dict) -> dict:
    if not sec_payload: return None
    ticker, f_type = sec_payload["ticker"], sec_payload["form_type"]
    prompt = f"""
    You are an elite equity research portfolio manager evaluating SEC disclosures ({f_type}) for {ticker}.
    1. Isolate CapEx ROI, supply chain bottlenecks, vendor concentrations, tariff/working-capital timing.
    2. Perform YoY Delta analysis on Item 1A Risk Factors.
    3. Analyze Recent 8-K Filings to extract material corporate events.
    4. Generate 3 highly targeted, institutional executive meeting questions.

    === MD&A ({f_type}) ===
    {sec_payload['mda']}
    === RISK FACTORS ({f_type}) ===
    {sec_payload['risk']}
    === PRIOR PERIOD RISK FACTORS ===
    {sec_payload['prior_risk']}
    === RECENT SEC FORM 8-K ===
    {sec_payload['eight_k']}
    """
    try:
        response = client.messages.create(
            model="claude-sonnet-5", max_tokens=4000,
            tools=[UNIFIED_EXTRACTION_TOOL], tool_choice={"type": "tool", "name": "record_comprehensive_firm_analysis"},
            messages=[{"role": "user", "content": prompt}]
        )
        
        for block in response.content:
            if block.type == "tool_use":
                raw_data = block.input
                bear_sorted = sorted(raw_data.get("bear_items", []), key=lambda x: {"HIGH": 0, "MEDIUM": 1}.get(x.get("impact", "MEDIUM").upper(), 2))
                bull_sorted = sorted(raw_data.get("bull_items", []), key=lambda x: {"HIGH": 0, "MEDIUM": 1}.get(x.get("impact", "MEDIUM").upper(), 2))
                high_threats = sum(1 for e in bear_sorted if e.get("impact") == "HIGH")
                return {
                    "ticker": ticker, "company_name": sec_payload["company_name"], "industry": sec_payload["industry"], "form_type": f_type,
                    "bear": bear_sorted, "bull": bull_sorted, "eight_k": raw_data.get("eight_k_insights", []),
                    "questions": raw_data.get("strategic_questions", []), "risk_score": min(10.0, round((high_threats * 1.5) + ((len(bear_sorted) - high_threats) * 0.5), 1)),
                    "financials": sec_payload["financials"], "insider_velocity": sec_payload["insider_velocity"], "event_markers": sec_payload["event_markers"]
                }
        return None
    except Exception as e: return None

def generate_interactive_chart(ticker, events):
    try:
        hist = yf.Ticker(ticker).history(period="6mo")
        if hist.empty: return None
        
        fig = go.Figure(data=[go.Candlestick(x=hist.index, open=hist['Open'], high=hist['High'], low=hist['Low'], close=hist['Close'], increasing_line_color='#089981', decreasing_line_color='#f23645', name='Price')])
        
        if events:
            e_dates, e_labels, y_vals = [], [], []
            for date_str, label in events:
                try:
                    dt = pd.to_datetime(date_str).tz_localize(hist.index.tz)
                    if dt in hist.index: y_val = hist.loc[dt]['High'] * 1.02
                    else: y_val = hist['High'].max()
                    e_dates.append(dt); e_labels.append(label); y_vals.append(y_val)
                except: pass
            
            if e_dates:
                fig.add_trace(go.Scatter(x=e_dates, y=y_vals, mode='markers+text', marker=dict(symbol='triangle-down', size=10, color='#2962ff'), text=e_labels, textposition='top center', textfont=dict(color='#2962ff', size=10, family="monospace"), name='SEC Event'))
                
        fig.update_layout(template="plotly_dark", margin=dict(l=0, r=0, t=10, b=0), height=300, paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)", xaxis_rangeslider_visible=False, showlegend=False, xaxis=dict(showgrid=False), yaxis=dict(showgrid=True, gridcolor='#2a2e39'))
        return fig
    except Exception: return None

def generate_excel(profiles):
    wb = openpyxl.Workbook()
    header_fill, header_font = PatternFill(start_color="1e222d", end_color="1e222d", fill_type="solid"), Font(color="ffffff", bold=True)
    
    ws_quant = wb.active
    ws_quant.title = "Financial Model Feed"
    ws_quant.append(['Ticker', 'Company Name', 'Gross Margin', 'Operating Margin', 'Net Margin', 'ROE', 'Debt to Equity', 'Free Cash Flow', 'EBITDA'])
    for p in profiles:
        ws_quant.append([p['ticker'], p['company_name'], p['financials']['gross_margin'], p['financials']['op_margin'], p['financials']['profit_margin'], p['financials']['roe'], p['financials']['debt_to_equity'], p['financials']['fcf'], p['financials']['ebitda']])

    ws_bull = wb.create_sheet(title="CapEx & Catalysts")
    ws_bull.append(['Ticker', 'Category', 'Target Initiative', 'Impact', 'Strategic Thesis', 'Filing Excerpt'])
    for p in profiles:
        for b in p['bull']:
            ws_bull.append([p['ticker'], b.get('category',''), b.get('target',''), b.get('impact',''), b.get('description',''), b.get('quote','')])

    ws_bear = wb.create_sheet(title="Supply Chain & Risks")
    ws_bear.append(['Ticker', 'Category', 'Target Risk', 'Impact', 'YoY Trend', 'Vulnerability', 'Filing Excerpt'])
    for p in profiles:
        for b in p['bear']:
            ws_bear.append([p['ticker'], b.get('category',''), b.get('target',''), b.get('impact',''), b.get('trend','STABLE'), b.get('description',''), b.get('quote','')])

    for ws in wb.worksheets:
        for cell in ws[1]: cell.fill, cell.font = header_fill, header_font
        for col in ws.columns: ws.column_dimensions[col[0].column_letter].width = 25
    
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()

# --- SLEEK COMPACT HEADER ---
c_head1, c_head2 = st.columns([3, 1])
c_head1.markdown("<h1>INSTITUTIONAL RESEARCH TERMINAL</h1>", unsafe_allow_html=True)
c_head1.markdown("<p style='color: #787b86; margin-top: -10px; font-size: 0.95rem;'>SEC EDGAR Ingestion • Form 4 Insider Velocity • Quant & Risk Analytics</p>", unsafe_allow_html=True)
c_head2.markdown("<div style='text-align: right; padding-top: 10px;'><span style='background: #1e222d; border: 1px solid #2a2e39; padding: 4px 10px; border-radius: 4px; font-family: monospace; font-size: 0.85rem; color: #089981;'>● LIVE FEED</span></div>", unsafe_allow_html=True)

st.divider()

# --- COMMAND TOOLBAR WITH INSTANT KEYUP SEARCH ---
if "tickers_input" not in st.session_state:
    st.session_state.tickers_input = "AAPL"

st.markdown("<div class='terminal-toolbar'>", unsafe_allow_html=True)
col_bar1, col_bar2, col_bar3, col_bar4 = st.columns([2, 1, 1, 1])

with col_bar1:
    tickers_input = st_keyup("Target Tickers (Comma-separated or Search)", value=st.session_state.tickers_input, placeholder="Search ticker or company name...", key="ticker_keyup_input", debounce=150)
with col_bar2:
    if st.button("Load Tech (AAPL)", use_container_width=True):
        st.session_state.tickers_input = "AAPL"
        st.rerun()
with col_bar3:
    if st.button("Load Industrials", use_container_width=True):
        st.session_state.tickers_input = "FAST, STRL"
        st.rerun()
with col_bar4:
    run_btn = st.button("⚡ Run Terminal", use_container_width=True)

st.markdown("</div>", unsafe_allow_html=True)

# --- REAL-TIME SMART MATCHING DROPDOWN (PREFIX & NAME PRIORITIZED) ---
sec_directory = load_sec_directory()
current_query = tickers_input.split(',')[-1].strip().upper()

if current_query and not sec_directory.empty:
    exact_ticker = sec_directory[sec_directory['ticker'].str.startswith(current_query, na=False)]
    exact_name = sec_directory[sec_directory['name'].str.contains(current_query, case=False, na=False) & ~sec_directory['ticker'].str.startswith(current_query, na=False)]
    matches = pd.concat([exact_ticker, exact_name]).head(5)
    
    if not matches.empty:
        st.markdown("<div class='autocomplete-dropdown'>", unsafe_allow_html=True)
        st.caption("🔍 Matching Ticker & Company Suggestions (Real-Time):")
        
        for idx, (_, row) in enumerate(matches.iterrows()):
            s_col1, s_col2 = st.columns([4, 1])
            with s_col1:
                st.markdown(f"""
                    <div style='padding: 6px 0;'>
                        <div style='font-size: 1.15rem; font-weight: 800; color: #ffffff; font-family: monospace;'>{row['ticker']}</div>
                        <div style='font-size: 0.88rem; color: #94a3b8;'>{row['name']}</div>
                    </div>
                """, unsafe_allow_html=True)
            with s_col2:
                if st.button("Select", key=f"sugg_btn_{row['ticker']}_{idx}", use_container_width=True):
                    parts = [t.strip() for t in tickers_input.split(',') if t.strip()]
                    if parts:
                        parts[-1] = row['ticker']
                    else:
                        parts = [row['ticker']]
                    st.session_state.tickers_input = ", ".join(parts)
                    st.rerun()
            if idx < len(matches) - 1:
                st.markdown("<hr style='margin: 4px 0; border-color: #2a2e39;'>", unsafe_allow_html=True)
                
        st.markdown("</div>", unsafe_allow_html=True)

# --- EXECUTION & HIGH-DENSITY DISPLAY ---
if run_btn:
    raw_list = [t.strip().upper() for t in tickers_input.split(',') if t.strip()]
    if raw_list:
        with st.spinner("Executing SEC EDGAR parsing, YoY risk delta, and quantitative modeling..."):
            profiles = []
            for t in raw_list:
                sec_data = fetch_firm_data(t)
                if sec_data:
                    prof = extract_unified_profile(sec_data)
                    if prof: profiles.append(prof)
            
            if profiles:
                st.success(f"Analysis successfully compiled for: {', '.join([p['ticker'] for p in profiles])}")
                st.divider()
                
                for p in profiles:
                    f = p['financials']
                    ins_vel = p['insider_velocity']
                    
                    st.markdown(f"### {p['ticker']} — {p['company_name']}")
                    st.caption(f"Industry: {p['industry']} | Filing: {p['form_type']}")
                    
                    cols = st.columns(6)
                    cols[0].metric("Insider 90D", str(ins_vel))
                    cols[1].metric("Price", f['price'])
                    cols[2].metric("Fwd P/E", f['fpe'])
                    cols[3].metric("Gross Mgn", f['gross_margin'])
                    cols[4].metric("FCF", f['fcf'])
                    cols[5].metric("Market Cap", f['mcap'])
                    
                    chart_fig = z = generate_interactive_chart(p['ticker'], p['event_markers'])
                    if chart_fig: st.plotly_chart(chart_fig, use_container_width=True)
                    
                    col1, col2 = st.columns(2)
                    with col1:
                        st.subheader("💡 CapEx & Growth Catalysts")
                        for item in p['bull']:
                            st.info(f"**{item['target']}** ({item['impact']})\n\n{item['description']}\n\n*{item['quote']}*")
                    with col2:
                        st.subheader("⚠️ Supply Chain & Risks")
                        for item in p['bear']:
                            st.warning(f"**{item['target']}** ({item['impact']} | {item.get('trend', 'STABLE')})\n\n{item['description']}\n\n*{item['quote']}*")
                    
                    st.subheader("🏛️ SEC Form 8-K Events")
                    if not p['eight_k']:
                        st.write("No recent material events found.")
                    else:
                        e_cols = st.columns(min(3, len(p['eight_k'])))
                        for idx, event in enumerate(p['eight_k']):
                            with e_cols[idx % 3]:
                                st.success(f"**{event['event_date']}** | {event['category']}\n\n{event['takeaway']}")
                    
                    st.subheader("🎙️ Executive Q&A / Meeting Prep")
                    for q in p['questions']: st.markdown(f"- {q}")
                    
                    st.markdown("<br>", unsafe_allow_html=True)
                    st.divider()

                excel_data = generate_excel(profiles)
                st.download_button(label="📊 Download Institutional Model (.xlsx)", data=excel_data, file_name="Institutional_Model.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
else:
    st.markdown("""
        <div style='background-color: #1e222d; border: 1px solid #2a2e39; padding: 20px; border-radius: 6px; text-align: center;'>
            <h3 style='color: #ffffff; margin-bottom: 8px;'>Terminal Ready</h3>
            <p style='color: #787b86; margin: 0;'>Select a preset basket above or type in the search bar to preview live stock suggestions, then click <b>Run Terminal</b>.</p>
        </div>
    """, unsafe_allow_html=True)
